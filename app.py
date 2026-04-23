from flask import Flask, render_template, request, redirect, url_for, session, Response, abort, send_from_directory, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import sqlite3
import os
import csv
import io
import secrets
import subprocess
import time
import re
import json
from collections import defaultdict
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
from weasyprint import HTML as WeasyprintHTML
import requests as http_requests
from bs4 import BeautifulSoup
from openai import OpenAI
from icalendar import Calendar, Event, vText
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

_login_attempts = defaultdict(list)
RATE_LIMIT_MAX = 5
RATE_LIMIT_WINDOW = 300

os.chdir(os.path.dirname(os.path.abspath(__file__)))

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'changez-cette-cle-en-production-svp')
if app.secret_key == 'changez-cette-cle-en-production-svp':
    print("⚠️  ATTENTION : SECRET_KEY par défaut utilisée. Changez-la en production !", flush=True)

_DATA_DIR = os.environ.get('DATA_DIR', '')
DB_NAME = os.path.join(_DATA_DIR, 'opensuivi.db') if _DATA_DIR else 'opensuivi.db'
UPLOAD_FOLDER = os.path.join(_DATA_DIR, 'uploads') if _DATA_DIR else os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')

STATUTS = [
    'Candidature envoyée',
    'Relance effectuée',
    'Entretien téléphonique',
    'Entretien présentiel',
    'Test technique',
    'Offre reçue',
    'Refus',
    'Désistement',
]

SECTEURS = [
    'Informatique / Tech',
    'Commerce / Vente',
    'Finance / Comptabilité',
    'Marketing / Communication',
    'RH / Formation',
    'Santé',
    'Industrie / Production',
    'Bâtiment / Travaux',
    'Transport / Logistique',
    'Autre',
]

TYPES_CONTRAT = ['Alternance', 'Stage', 'CDI', 'CDD', 'Freelance', 'Intérim', 'Autre']

PLATEFORMES = ['LinkedIn', 'Indeed', 'France Travail', 'APEC', 'HelloWork', 'Welcome to the Jungle', 'Direct', 'Réseau', 'Autre']

TYPES_CONTACT = [
    'Appel téléphonique',
    'Email envoyé',
    'Email reçu',
    'Entretien téléphonique',
    'Entretien visio',
    'Entretien présentiel',
    'Test technique',
    'Note',
    'Autre',
]

def init_db():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS prospects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        etablissement TEXT NOT NULL,
        contact TEXT,
        telephone TEXT,
        email TEXT,
        adresse TEXT,
        code_postal TEXT DEFAULT '',
        ville TEXT DEFAULT '',
        categorie TEXT,
        statut TEXT,
        commentaire TEXT,
        date_relance DATE,
        date_ajout TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        archived INTEGER DEFAULT 0
    )''')
    for col_def in [
        ('poste',         'TEXT DEFAULT ""'),
        ('type_contrat',  'TEXT DEFAULT ""'),
        ('plateforme',    'TEXT DEFAULT ""'),
        ('cv_envoye',     'INTEGER DEFAULT 0'),
        ('lettre_envoyee','INTEGER DEFAULT 0'),
        ('code_postal',   'TEXT DEFAULT ""'),
        ('ville',         'TEXT DEFAULT ""'),
        ('archived',      'INTEGER DEFAULT 0'),
        ('lien_offre',             'TEXT DEFAULT ""'),
        ('mode_candidature',       'TEXT DEFAULT ""'),
        ('reference_offre',        'TEXT DEFAULT ""'),
        ('date_limite_candidature','DATE DEFAULT NULL'),
        ('date_debut',             'DATE DEFAULT NULL'),
        ('duree_contrat',          'TEXT DEFAULT ""'),
        ('formation_requise',      'TEXT DEFAULT ""'),
        # anciens champs conservés pour compatibilité DB, ignorés dans l'UI
        ('temperature',   'TEXT DEFAULT ""'),
        ('catalogue',     'INTEGER DEFAULT 0'),
        ('grille',        'INTEGER DEFAULT 0'),
        ('devis',         'INTEGER DEFAULT 0'),
        ('source',        'TEXT DEFAULT ""'),
        ('base_legale',   'TEXT DEFAULT ""'),
        ('no_contact',    'INTEGER DEFAULT 0'),
    ]:
        try:
            c.execute(f"ALTER TABLE prospects ADD COLUMN {col_def[0]} {col_def[1]}")
        except sqlite3.OperationalError:
            pass

    c.execute('''CREATE TABLE IF NOT EXISTS interventions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        prospect_id INTEGER,
        date_intervention TEXT,
        type_contact TEXT,
        compte_rendu TEXT,
        username TEXT DEFAULT '',
        FOREIGN KEY(prospect_id) REFERENCES prospects(id)
    )''')
    try:
        c.execute("ALTER TABLE interventions ADD COLUMN username TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL
    )''')
    try:
        c.execute("ALTER TABLE users ADD COLUMN signature TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass

    c.execute('''CREATE TABLE IF NOT EXISTS evenements_perso (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titre TEXT NOT NULL,
        date_event DATE NOT NULL,
        type_event TEXT NOT NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS logs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        action TEXT,
        ip TEXT,
        horodatage TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS historique_statut (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        prospect_id INTEGER,
        ancien_statut TEXT,
        nouveau_statut TEXT,
        username TEXT,
        horodatage TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(prospect_id) REFERENCES prospects(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS tags (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT UNIQUE NOT NULL
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS prospect_tags (
        prospect_id INTEGER,
        tag_id INTEGER,
        PRIMARY KEY(prospect_id, tag_id),
        FOREIGN KEY(prospect_id) REFERENCES prospects(id),
        FOREIGN KEY(tag_id) REFERENCES tags(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS intervention_attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        intervention_id INTEGER,
        filename TEXT,
        original_name TEXT,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(intervention_id) REFERENCES interventions(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS prospect_documents (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        prospect_id INTEGER NOT NULL,
        doc_type TEXT NOT NULL,
        filename TEXT NOT NULL,
        original_name TEXT NOT NULL,
        uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(prospect_id) REFERENCES prospects(id)
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS user_settings (
        user_id INTEGER,
        key TEXT,
        value TEXT,
        PRIMARY KEY (user_id, key),
        FOREIGN KEY (user_id) REFERENCES users(id)
    )''')

    # Colonnes multi-utilisateurs + nouvelles fonctionnalités
    for tbl, col, defval in [
        ('users',          'approved',         'INTEGER DEFAULT 1'),
        ('users',          'ical_token',        'TEXT DEFAULT NULL'),
        ('prospects',      'user_id',           'INTEGER DEFAULT NULL'),
        ('prospects',      'notes_entretien',   'TEXT DEFAULT NULL'),
        ('evenements_perso','user_id',          'INTEGER DEFAULT NULL'),
    ]:
        try:
            c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} {defval}")
        except sqlite3.OperationalError:
            pass

    # Générer un ical_token pour les users qui n'en ont pas encore
    rows = c.execute("SELECT id FROM users WHERE ical_token IS NULL").fetchall()
    for row in rows:
        c.execute("UPDATE users SET ical_token=? WHERE id=?", (str(uuid.uuid4()), row[0]))

    default_app_name = os.environ.get('APP_NAME', 'Suivi Emploi')
    c.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('company_name', ?)", (default_app_name,))

    c.execute("SELECT * FROM users WHERE username = 'admin'")
    if not c.fetchone():
        hashed_pw = generate_password_hash('admin123')
        c.execute("INSERT INTO users (username, password, role, approved) VALUES (?, ?, ?, 1)", ('admin', hashed_pw, 'admin'))

    # S'assurer que l'admin est approuvé
    c.execute("UPDATE users SET approved=1 WHERE role='admin'")

    # Migrer les données existantes vers l'admin
    c.execute("SELECT id FROM users WHERE role='admin' LIMIT 1")
    admin_row = c.fetchone()
    if admin_row:
        admin_id = admin_row[0]
        c.execute("UPDATE prospects SET user_id=? WHERE user_id IS NULL", (admin_id,))
        c.execute("UPDATE evenements_perso SET user_id=? WHERE user_id IS NULL", (admin_id,))
        # Migrer profil_recherche global → user_settings de l'admin
        for key in ['profil_type_contrat','profil_niveau','profil_specialite','profil_ecole',
                    'profil_rythme','profil_dispo','profil_localisation','profil_salaire']:
            row = c.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
            if row and row[0]:
                c.execute("INSERT OR IGNORE INTO user_settings (user_id, key, value) VALUES (?,?,?)",
                          (admin_id, key, row[0]))

    conn.commit()
    conn.close()

def get_setting(key, default=''):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT value FROM settings WHERE key = ?", (key,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else default

def get_openai_key():
    return get_setting('openai_api_key', '')

def get_user_setting(user_id, key, default=''):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT value FROM user_settings WHERE user_id=? AND key=?", (user_id, key))
    row = c.fetchone()
    conn.close()
    return row[0] if row else default

def set_user_setting(user_id, key, value):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO user_settings (user_id, key, value) VALUES (?,?,?)", (user_id, key, value))
    conn.commit()
    conn.close()

def _check_prospect_owner(c, prospect_id, user_id):
    c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (prospect_id, user_id))
    if not c.fetchone():
        abort(403)

def get_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']

def validate_csrf():
    token = request.form.get('csrf_token')
    if not token or token != session.get('csrf_token'):
        abort(403)

@app.template_filter('tel')
def format_tel(value):
    if not value:
        return ''
    digits = ''.join(c for c in str(value) if c.isdigit())[:10]
    return ' '.join(digits[i:i+2] for i in range(0, len(digits), 2))

@app.before_request
def csrf_protect():
    if request.method == 'POST' and request.endpoint != 'login':
        if request.is_json:
            data = request.get_json(silent=True) or {}
            token = data.get('csrf_token')
            if not token or token != session.get('csrf_token'):
                abort(403)
        else:
            validate_csrf()

@app.context_processor
def inject_globals():
    logo_filename = get_setting('logo_filename', '')
    logo_url = f'/static/{logo_filename}' if logo_filename else ''
    custom_colors = {
        'primary': get_setting('color_primary', ''),
        'accent':  get_setting('color_accent', ''),
    }
    uid = session.get('user_id')
    if uid:
        profil = {
            'type_contrat': get_user_setting(uid, 'profil_type_contrat', ''),
            'niveau':       get_user_setting(uid, 'profil_niveau', ''),
            'specialite':   get_user_setting(uid, 'profil_specialite', ''),
            'ecole':        get_user_setting(uid, 'profil_ecole', ''),
            'rythme':       get_user_setting(uid, 'profil_rythme', ''),
            'dispo':        get_user_setting(uid, 'profil_dispo', ''),
            'localisation': get_user_setting(uid, 'profil_localisation', ''),
            'salaire':      get_user_setting(uid, 'profil_salaire', ''),
        }
        try:
            conn_nb = sqlite3.connect(DB_NAME)
            nb_row = conn_nb.execute(
                "SELECT COUNT(*) FROM prospects WHERE date_relance <= date('now') "
                "AND statut NOT IN ('Refus','Désistement','Offre reçue') "
                "AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,)
            ).fetchone()
            nb_relances_retard = nb_row[0] if nb_row else 0
            conn_nb.close()
        except Exception:
            nb_relances_retard = 0
    else:
        profil = {k: '' for k in ['type_contrat','niveau','specialite','ecole','rythme','dispo','localisation','salaire']}
        nb_relances_retard = 0
    return dict(
        theme=session.get('theme', 'light'),
        csrf_token=get_csrf_token(),
        company_name=get_setting('company_name', 'Suivi Emploi'),
        logo_url=logo_url,
        custom_colors=custom_colors,
        profil_recherche=profil,
        STATUTS=STATUTS,
        SECTEURS=SECTEURS,
        TYPES_CONTRAT=TYPES_CONTRAT,
        PLATEFORMES=PLATEFORMES,
        TYPES_CONTACT=TYPES_CONTACT,
        nb_relances_retard=nb_relances_retard,
    )

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ─────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        now = time.time()
        _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < RATE_LIMIT_WINDOW]
        if len(_login_attempts[ip]) >= RATE_LIMIT_MAX:
            remaining = int(RATE_LIMIT_WINDOW - (now - _login_attempts[ip][0]))
            error = f"Trop de tentatives. Réessayez dans {remaining // 60 + 1} minute(s)."
            return render_template('login.html', error=error)
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if not username or not password:
            error = "Veuillez remplir tous les champs."
        else:
            conn = sqlite3.connect(DB_NAME)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT * FROM users WHERE username = ?", (username,))
            user = c.fetchone()
            conn.close()
            if user and check_password_hash(user['password'], password):
                if not user['approved']:
                    error = "Votre compte est en attente de validation par l'administrateur."
                else:
                    _login_attempts[ip] = []
                    session['user_id'] = user['id']
                    session['username'] = user['username']
                    session['role'] = user['role']
                    conn2 = sqlite3.connect(DB_NAME); c2 = conn2.cursor()
                    c2.execute("INSERT INTO logs (username, action, ip) VALUES (?, ?, ?)", (username, 'connexion', ip))
                    conn2.commit(); conn2.close()
                    return redirect(url_for('index'))
            else:
                _login_attempts[ip].append(now)
                remaining_attempts = RATE_LIMIT_MAX - len(_login_attempts[ip])
                error = f"Identifiants incorrects.{f' ({remaining_attempts} tentative(s) restante(s))' if remaining_attempts < RATE_LIMIT_MAX else ''}"
    return render_template('login.html', error=error)

@app.route('/inscription', methods=['GET', 'POST'])
def inscription():
    if 'user_id' in session:
        return redirect(url_for('index'))
    error = None
    success = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')
        if not username or not password:
            error = "Nom d'utilisateur et mot de passe obligatoires."
        elif len(username) < 3:
            error = "Le nom d'utilisateur doit faire au moins 3 caractères."
        elif len(password) < 6:
            error = "Le mot de passe doit faire au moins 6 caractères."
        elif password != confirm:
            error = "Les mots de passe ne correspondent pas."
        else:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("SELECT id FROM users WHERE username=?", (username,))
            if c.fetchone():
                error = "Ce nom d'utilisateur est déjà pris."
                conn.close()
            else:
                hashed_pw = generate_password_hash(password)
                c.execute("INSERT INTO users (username, password, role, approved) VALUES (?,?,?,0)",
                          (username, hashed_pw, 'commercial'))
                conn.commit()
                conn.close()
                success = True
    return render_template('inscription.html', error=error, success=success)

@app.route('/logout')
def logout():
    if 'username' in session:
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        c.execute("INSERT INTO logs (username, action, ip) VALUES (?, ?, ?)", (session['username'], 'déconnexion', ip))
        conn.commit(); conn.close()
    session.clear()
    return redirect(url_for('login'))

@app.route('/toggle-theme')
@login_required
def toggle_theme():
    current_theme = session.get('theme', 'light')
    session['theme'] = 'dark' if current_theme == 'light' else 'light'
    return redirect(request.referrer or url_for('index'))

# ─────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    uid = session['user_id']
    c.execute("SELECT * FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=? ORDER BY date_ajout DESC", (uid,))
    prospects = c.fetchall()
    total = len(prospects)
    cv_envoyes = sum(1 for p in prospects if p['cv_envoye'] == 1)
    entretiens = sum(1 for p in prospects if p['statut'] in ('Entretien téléphonique', 'Entretien présentiel', 'Test technique'))
    offres = sum(1 for p in prospects if p['statut'] == 'Offre reçue')
    refus = sum(1 for p in prospects if p['statut'] == 'Refus')
    taux_reponse = round((entretiens / total * 100) if total > 0 else 0, 1)

    aujourdhui = datetime.now().strftime('%Y-%m-%d')
    debut_mois = datetime.now().strftime('%Y-%m-01')

    a_relancer = [p for p in prospects if p['date_relance'] and p['date_relance'] <= aujourdhui
                  and p['statut'] not in ('Offre reçue', 'Refus', 'Désistement')]
    relances_auj = [p for p in a_relancer if p['date_relance'] == aujourdhui]

    c.execute("SELECT COUNT(*) FROM prospects WHERE date_ajout >= ? AND (archived=0 OR archived IS NULL) AND user_id=?", (debut_mois, uid))
    nouvelles_mois = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM prospects WHERE id NOT IN (SELECT DISTINCT prospect_id FROM interventions) AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    sans_suivi = c.fetchone()[0]

    il_y_a_7j = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    c.execute("""SELECT i.*, p.etablissement, p.poste FROM interventions i
                 JOIN prospects p ON p.id = i.prospect_id
                 WHERE i.date_intervention >= ? AND p.user_id=? ORDER BY i.id DESC LIMIT 10""", (il_y_a_7j, uid))
    activite_recente = c.fetchall()

    stats_statuts = {}
    for p in prospects:
        s = p['statut'] or 'Non défini'
        stats_statuts[s] = stats_statuts.get(s, 0) + 1

    c.execute("""SELECT categorie, COUNT(*) as total
                 FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=?
                 AND categorie IS NOT NULL AND categorie != ''
                 GROUP BY categorie ORDER BY total DESC""", (uid,))
    stats_secteurs = [dict(r) for r in c.fetchall()]

    conn.close()
    return render_template('index.html',
        total=total, cv_envoyes=cv_envoyes, entretiens=entretiens,
        offres=offres, refus=refus, taux_reponse=taux_reponse,
        a_relancer=a_relancer, relances_auj=relances_auj,
        nouvelles_mois=nouvelles_mois, sans_suivi=sans_suivi,
        activite_recente=activite_recente,
        stats_statuts=stats_statuts, stats_secteurs=stats_secteurs)

# ─────────────────────────────────────────────
# CANDIDATURES
# ─────────────────────────────────────────────

@app.route('/base')
@login_required
def base_donnees():
    q = request.args.get('q', '').strip()
    f_statut = request.args.get('statut', '').strip()
    f_secteur = request.args.get('secteur', '').strip()
    f_type_contrat = request.args.get('type_contrat', '').strip()
    f_ville = request.args.get('ville', '').strip()
    f_tag = request.args.get('tag', '').strip()
    sort = request.args.get('sort', 'date_ajout').strip()
    order = request.args.get('order', 'desc').strip()
    try:
        page = max(1, int(request.args.get('page', 1) or 1))
    except (ValueError, TypeError):
        page = 1
    per_page = 50

    sort_cols = {
        'etablissement': 'p.etablissement',
        'poste': 'p.poste',
        'statut': 'p.statut',
        'date_relance': 'p.date_relance',
        'date_ajout': 'p.date_ajout',
        'derniere_activite': 'derniere_activite',
    }
    sort_col = sort_cols.get(sort, 'p.date_ajout')
    order_sql = 'ASC' if order == 'asc' else 'DESC'

    uid = session['user_id']
    conditions = ["(p.archived IS NULL OR p.archived = 0)", "p.user_id = ?"]
    params = [uid]
    if q:
        conditions.append("(p.etablissement LIKE ? OR p.contact LIKE ? OR p.poste LIKE ? OR p.email LIKE ? OR p.ville LIKE ?)")
        like = f'%{q}%'
        params += [like, like, like, like, like]
    if f_statut:
        conditions.append("p.statut = ?")
        params.append(f_statut)
    if f_secteur:
        conditions.append("p.categorie = ?")
        params.append(f_secteur)
    if f_type_contrat:
        conditions.append("p.type_contrat = ?")
        params.append(f_type_contrat)
    if f_ville:
        conditions.append("p.ville LIKE ?")
        params.append(f'%{f_ville}%')
    if f_tag:
        conditions.append("EXISTS (SELECT 1 FROM prospect_tags pt JOIN tags t ON t.id=pt.tag_id WHERE pt.prospect_id=p.id AND t.nom=?)")
        params.append(f_tag)

    where = "WHERE " + " AND ".join(conditions)

    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute(f"SELECT COUNT(DISTINCT p.id) FROM prospects p LEFT JOIN interventions i ON i.prospect_id = p.id {where}", params)
    total_count = c.fetchone()[0]

    offset = (page - 1) * per_page
    c.execute(f"""SELECT p.*, MAX(i.date_intervention) as derniere_activite
                  FROM prospects p
                  LEFT JOIN interventions i ON i.prospect_id = p.id
                  {where}
                  GROUP BY p.id
                  ORDER BY {sort_col} {order_sql} NULLS LAST
                  LIMIT ? OFFSET ?""", params + [per_page, offset])
    prospects = c.fetchall()

    c.execute("SELECT DISTINCT statut FROM prospects WHERE user_id=? AND statut IS NOT NULL AND statut != '' ORDER BY statut", (uid,))
    statuts = [r[0] for r in c.fetchall()]
    c.execute("SELECT DISTINCT categorie FROM prospects WHERE user_id=? AND categorie IS NOT NULL AND categorie != '' ORDER BY categorie", (uid,))
    secteurs = [r[0] for r in c.fetchall()]
    c.execute("SELECT DISTINCT ville FROM prospects WHERE user_id=? AND ville IS NOT NULL AND ville != '' ORDER BY ville", (uid,))
    villes = [r[0] for r in c.fetchall()]
    c.execute("SELECT nom FROM tags ORDER BY nom ASC")
    all_tags = [r[0] for r in c.fetchall()]

    conn.close()
    total_pages = max(1, (total_count + per_page - 1) // per_page)

    today = datetime.now().date()
    today_str = today.strftime('%Y-%m-%d')
    today_plus3 = (today + timedelta(days=3)).strftime('%Y-%m-%d')
    return render_template('base_donnees.html', prospects=prospects,
                           total_count=total_count, page=page, total_pages=total_pages,
                           statuts=statuts, secteurs=secteurs, villes=villes, all_tags=all_tags,
                           q=q, f_statut=f_statut, f_secteur=f_secteur,
                           f_type_contrat=f_type_contrat, f_ville=f_ville, f_tag=f_tag,
                           sort=sort, order=order,
                           today_str=today_str, today_plus3=today_plus3)

@app.route('/nouveau')
@login_required
def nouveau_prospect_form():
    return render_template('nouveau.html')

@app.route('/ajouter', methods=['POST'])
@login_required
def ajouter_prospect():
    etablissement = request.form.get('etablissement', '').strip()
    if not etablissement:
        return redirect(url_for('nouveau_prospect_form'))
    contact = request.form.get('contact', '').strip()
    telephone = request.form.get('telephone', '').strip()
    email = request.form.get('email', '').strip()
    adresse = request.form.get('adresse', '').strip()
    code_postal = request.form.get('code_postal', '').strip()
    ville = request.form.get('ville', '').strip()
    categorie = request.form.get('categorie', '').strip()
    poste = request.form.get('poste', '').strip()
    type_contrat = request.form.get('type_contrat', '').strip()
    plateforme = request.form.get('plateforme', '').strip()
    statut = request.form.get('statut', 'Candidature envoyée').strip()
    date_relance = request.form.get('date_relance', '').strip() or None
    commentaire = request.form.get('commentaire', '').strip()
    cv_envoye = 1 if 'cv_envoye' in request.form else 0
    lettre_envoyee = 1 if 'lettre_envoyee' in request.form else 0
    lien_offre = request.form.get('lien_offre', '').strip()
    mode_candidature = request.form.get('mode_candidature', '').strip()
    reference_offre = request.form.get('reference_offre', '').strip()
    date_limite_candidature = request.form.get('date_limite_candidature', '').strip() or None
    date_debut = request.form.get('date_debut', '').strip() or None
    duree_contrat = request.form.get('duree_contrat', '').strip()
    formation_requise = request.form.get('formation_requise', '').strip()
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute(
        """INSERT INTO prospects
           (etablissement, contact, telephone, email, adresse, code_postal, ville,
            categorie, poste, type_contrat, plateforme, statut, date_relance,
            commentaire, cv_envoye, lettre_envoyee,
            lien_offre, mode_candidature, reference_offre,
            date_limite_candidature, date_debut, duree_contrat, formation_requise,
            user_id)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (etablissement, contact, telephone, email, adresse, code_postal, ville,
         categorie, poste, type_contrat, plateforme, statut, date_relance,
         commentaire, cv_envoye, lettre_envoyee,
         lien_offre, mode_candidature, reference_offre,
         date_limite_candidature, date_debut, duree_contrat, formation_requise,
         session['user_id'])
    )
    conn.commit()
    conn.close()
    return redirect(url_for('base_donnees'))

@app.route('/prospect/<int:id>')
@login_required
def voir_prospect(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    prospect = c.fetchone()
    if not prospect:
        conn.close()
        return redirect(url_for('base_donnees'))
    c.execute("SELECT * FROM interventions WHERE prospect_id = ? ORDER BY date_intervention DESC", (id,))
    interventions = c.fetchall()
    c.execute("SELECT * FROM historique_statut WHERE prospect_id = ? ORDER BY horodatage DESC", (id,))
    historique_statut = c.fetchall()
    c.execute("SELECT t.nom FROM tags t JOIN prospect_tags pt ON pt.tag_id=t.id WHERE pt.prospect_id=? ORDER BY t.nom", (id,))
    prospect_tags = [r[0] for r in c.fetchall()]
    intervention_ids = [i['id'] for i in interventions]
    attachments_map = {}
    if intervention_ids:
        placeholders = ','.join('?' for _ in intervention_ids)
        c.execute(f"SELECT * FROM intervention_attachments WHERE intervention_id IN ({placeholders})", intervention_ids)
        for att in c.fetchall():
            attachments_map.setdefault(att['intervention_id'], []).append(att)
    docs = {r['doc_type']: r for r in c.execute(
        "SELECT * FROM prospect_documents WHERE prospect_id=?", (id,)).fetchall()}
    conn.close()
    return render_template('prospect.html', prospect=prospect, interventions=interventions,
                           historique_statut=historique_statut,
                           prospect_tags=prospect_tags, attachments_map=attachments_map,
                           prospect_docs=docs,
                           get_openai_key_set=bool(get_openai_key()),
                           now_date=datetime.now().strftime('%Y-%m-%d'))

@app.route('/prospect/<int:id>/intervention', methods=['POST'])
@login_required
def ajouter_intervention(id):
    date_intervention = request.form.get('date_intervention', '').strip()
    type_contact = request.form.get('type_contact', '').strip()
    compte_rendu = request.form.get('compte_rendu', '').strip()
    nouveau_statut = request.form.get('nouveau_statut', '').strip()
    nouvelle_relance = request.form.get('nouvelle_relance', '').strip() or None
    cv_envoye = 1 if 'cv_envoye' in request.form else None
    lettre_envoyee = 1 if 'lettre_envoyee' in request.form else None
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    _check_prospect_owner(c, id, session['user_id'])
    c.execute("SELECT statut FROM prospects WHERE id = ?", (id,))
    row = c.fetchone()
    ancien_statut = row['statut'] if row else ''
    c.execute(
        "INSERT INTO interventions (prospect_id, date_intervention, type_contact, compte_rendu, username) VALUES (?, ?, ?, ?, ?)",
        (id, date_intervention, type_contact, compte_rendu, session.get('username', ''))
    )
    updates = ['statut=?', 'date_relance=?']
    vals = [nouveau_statut, nouvelle_relance]
    if cv_envoye is not None:
        updates.append('cv_envoye=?')
        vals.append(cv_envoye)
    if lettre_envoyee is not None:
        updates.append('lettre_envoyee=?')
        vals.append(lettre_envoyee)
    vals.append(id)
    c.execute(f"UPDATE prospects SET {', '.join(updates)} WHERE id=?", vals)
    if nouveau_statut and nouveau_statut != ancien_statut:
        c.execute(
            "INSERT INTO historique_statut (prospect_id, ancien_statut, nouveau_statut, username) VALUES (?, ?, ?, ?)",
            (id, ancien_statut, nouveau_statut, session.get('username'))
        )
    conn.commit()
    conn.close()
    return redirect(url_for('voir_prospect', id=id))

@app.route('/prospect/<int:id>/editer', methods=['GET', 'POST'])
@login_required
def editer_prospect(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    _check_prospect_owner(c, id, session['user_id'])
    if request.method == 'POST':
        etablissement = request.form.get('etablissement', '').strip()
        if not etablissement:
            c.execute("SELECT * FROM prospects WHERE id = ?", (id,))
            prospect = c.fetchone()
            conn.close()
            return render_template('editer_prospect.html', prospect=prospect, error="L'entreprise est obligatoire.")
        contact = request.form.get('contact', '').strip()
        telephone = request.form.get('telephone', '').strip()
        email = request.form.get('email', '').strip()
        adresse = request.form.get('adresse', '').strip()
        code_postal = request.form.get('code_postal', '').strip()
        ville = request.form.get('ville', '').strip()
        categorie = request.form.get('categorie', '').strip()
        poste = request.form.get('poste', '').strip()
        type_contrat = request.form.get('type_contrat', '').strip()
        plateforme = request.form.get('plateforme', '').strip()
        statut = request.form.get('statut', '').strip()
        date_relance = request.form.get('date_relance', '').strip() or None
        commentaire = request.form.get('commentaire', '').strip()
        cv_envoye = 1 if 'cv_envoye' in request.form else 0
        lettre_envoyee = 1 if 'lettre_envoyee' in request.form else 0
        c.execute("SELECT statut FROM prospects WHERE id = ?", (id,))
        row_statut = c.fetchone()
        ancien_statut = row_statut['statut'] if row_statut else ''
        c.execute(
            '''UPDATE prospects SET etablissement=?, contact=?, telephone=?, email=?, adresse=?,
               code_postal=?, ville=?, categorie=?, poste=?, type_contrat=?, plateforme=?,
               statut=?, date_relance=?, commentaire=?, cv_envoye=?, lettre_envoyee=? WHERE id=?''',
            (etablissement, contact, telephone, email, adresse, code_postal, ville,
             categorie, poste, type_contrat, plateforme, statut, date_relance,
             commentaire, cv_envoye, lettre_envoyee, id)
        )
        if statut and statut != ancien_statut:
            c.execute(
                "INSERT INTO historique_statut (prospect_id, ancien_statut, nouveau_statut, username) VALUES (?, ?, ?, ?)",
                (id, ancien_statut, statut, session.get('username'))
            )
        conn.commit()
        conn.close()
        return redirect(url_for('voir_prospect', id=id))
    c.execute("SELECT * FROM prospects WHERE id = ?", (id,))
    prospect = c.fetchone()
    conn.close()
    if not prospect:
        return redirect(url_for('base_donnees'))
    return render_template('editer_prospect.html', prospect=prospect)

@app.route('/prospect/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_prospect(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    _check_prospect_owner(c, id, session['user_id'])
    c.execute("SELECT etablissement FROM prospects WHERE id = ?", (id,))
    row = c.fetchone()
    nom = row['etablissement'] if row else f"ID#{id}"
    c.execute("DELETE FROM interventions WHERE prospect_id = ?", (id,))
    c.execute("DELETE FROM historique_statut WHERE prospect_id = ?", (id,))
    c.execute("DELETE FROM prospect_tags WHERE prospect_id = ?", (id,))
    c.execute("DELETE FROM prospects WHERE id = ?", (id,))
    c.execute("INSERT INTO logs (username, action, ip) VALUES (?, ?, ?)",
              (session.get('username'), f"Suppression candidature : {nom} (ID#{id})", request.remote_addr))
    conn.commit()
    conn.close()
    return redirect(url_for('base_donnees'))

@app.route('/prospect/<int:id>/archiver', methods=['POST'])
@login_required
def archiver_prospect(id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    _check_prospect_owner(c, id, session['user_id'])
    c.execute("UPDATE prospects SET archived = 1 WHERE id = ?", (id,))
    c.execute("INSERT INTO logs (username, action, ip) VALUES (?, ?, ?)",
              (session.get('username'), f"Archivage candidature ID#{id}", request.remote_addr))
    conn.commit()
    conn.close()
    return redirect(url_for('base_donnees'))

@app.route('/prospect/<int:id>/restaurer', methods=['POST'])
@login_required
def restaurer_prospect(id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    _check_prospect_owner(c, id, session['user_id'])
    c.execute("UPDATE prospects SET archived = 0 WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('corbeille'))

@app.route('/corbeille')
@login_required
def corbeille():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE archived=1 AND user_id=? ORDER BY etablissement ASC", (session['user_id'],))
    prospects = c.fetchall()
    conn.close()
    return render_template('corbeille.html', prospects=prospects)

# ─────────────────────────────────────────────
# TAGS
# ─────────────────────────────────────────────

@app.route('/api/tags')
@login_required
def api_tags():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT nom FROM tags ORDER BY nom ASC")
    tags = [r[0] for r in c.fetchall()]
    conn.close()
    return jsonify(tags)

@app.route('/prospect/<int:id>/tags', methods=['POST'])
@login_required
def update_prospect_tags(id):
    data = request.get_json(silent=True) or {}
    tags_list = [t.strip() for t in data.get('tags', []) if t.strip()]
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    for tag in tags_list:
        c.execute("INSERT OR IGNORE INTO tags (nom) VALUES (?)", (tag,))
    c.execute("DELETE FROM prospect_tags WHERE prospect_id = ?", (id,))
    for tag in tags_list:
        c.execute("SELECT id FROM tags WHERE nom = ?", (tag,))
        row = c.fetchone()
        if row:
            c.execute("INSERT OR IGNORE INTO prospect_tags (prospect_id, tag_id) VALUES (?, ?)", (id, row[0]))
    conn.commit()
    conn.close()
    return jsonify(ok=True, tags=tags_list)

# ─────────────────────────────────────────────
# PIÈCES JOINTES
# ─────────────────────────────────────────────

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/intervention/<int:id>/attachment', methods=['POST'])
@login_required
def upload_attachment(id):
    if 'file' not in request.files:
        return redirect(request.referrer or url_for('base_donnees'))
    f = request.files['file']
    if not f or not f.filename or not allowed_file(f.filename):
        return redirect(request.referrer or url_for('base_donnees'))
    original_name = f.filename
    ext = original_name.rsplit('.', 1)[1].lower()
    unique_name = f"{secrets.token_hex(12)}.{ext}"
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    f.save(os.path.join(UPLOAD_FOLDER, unique_name))
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("INSERT INTO intervention_attachments (intervention_id, filename, original_name) VALUES (?, ?, ?)",
              (id, unique_name, original_name))
    c.execute("SELECT prospect_id FROM interventions WHERE id = ?", (id,))
    row = c.fetchone()
    prospect_id = row['prospect_id'] if row else None
    conn.commit()
    conn.close()
    if prospect_id:
        return redirect(url_for('voir_prospect', id=prospect_id))
    return redirect(url_for('base_donnees'))

@app.route('/attachment/<int:id>')
@login_required
def download_attachment(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM intervention_attachments WHERE id = ?", (id,))
    att = c.fetchone()
    conn.close()
    if not att:
        abort(404)
    return send_from_directory(UPLOAD_FOLDER, att['filename'],
                               as_attachment=True, download_name=att['original_name'])

@app.route('/attachment/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_attachment(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT ia.*, i.prospect_id FROM intervention_attachments ia JOIN interventions i ON i.id=ia.intervention_id WHERE ia.id=?", (id,))
    att = c.fetchone()
    if att:
        filepath = os.path.join(UPLOAD_FOLDER, att['filename'])
        if os.path.exists(filepath):
            os.remove(filepath)
        c.execute("DELETE FROM intervention_attachments WHERE id = ?", (id,))
        prospect_id = att['prospect_id']
    conn.commit()
    conn.close()
    if att and prospect_id:
        return redirect(url_for('voir_prospect', id=prospect_id))
    return redirect(url_for('base_donnees'))

# ─────────────────────────────────────────────
# EXPORT PDF
# ─────────────────────────────────────────────

@app.route('/prospect/<int:id>/pdf')
@login_required
def export_prospect_pdf(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE id = ?", (id,))
    prospect = c.fetchone()
    if not prospect:
        conn.close()
        return "Candidature introuvable", 404
    c.execute("SELECT * FROM interventions WHERE prospect_id = ? ORDER BY date_intervention DESC", (id,))
    interventions = c.fetchall()
    c.execute("SELECT * FROM historique_statut WHERE prospect_id = ? ORDER BY horodatage DESC", (id,))
    historique_statut = c.fetchall()
    conn.close()
    profil = {
        'type_contrat': get_setting('profil_type_contrat', ''),
        'niveau':       get_setting('profil_niveau', ''),
        'specialite':   get_setting('profil_specialite', ''),
        'ecole':        get_setting('profil_ecole', ''),
        'dispo':        get_setting('profil_dispo', ''),
        'localisation': get_setting('profil_localisation', ''),
    }
    html_string = render_template('prospect_pdf.html', prospect=prospect,
                                  interventions=interventions, historique_statut=historique_statut,
                                  profil=profil,
                                  now=datetime.now().strftime('%d/%m/%Y %H:%M'))
    pdf_bytes = WeasyprintHTML(string=html_string, base_url=request.base_url).write_pdf()
    safe_name = (prospect['etablissement'] or 'candidature')[:30].replace('/', '-')
    return Response(pdf_bytes, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename=candidature_{id}_{safe_name}.pdf'})

# ─────────────────────────────────────────────
# INTERVENTIONS
# ─────────────────────────────────────────────

@app.route('/intervention/<int:id>/editer', methods=['GET', 'POST'])
@login_required
def editer_intervention(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM interventions WHERE id = ?", (id,))
    intervention = c.fetchone()
    if not intervention:
        conn.close()
        return redirect(url_for('base_donnees'))
    prospect_id = intervention['prospect_id']
    if request.method == 'POST':
        date_intervention = request.form.get('date_intervention', '').strip()
        type_contact = request.form.get('type_contact', '').strip()
        compte_rendu = request.form.get('compte_rendu', '').strip()
        nouveau_statut = request.form.get('nouveau_statut', '').strip()
        nouvelle_relance = request.form.get('nouvelle_relance', '').strip() or None
        c.execute(
            'UPDATE interventions SET date_intervention=?, type_contact=?, compte_rendu=? WHERE id=?',
            (date_intervention, type_contact, compte_rendu, id)
        )
        c.execute(
            'UPDATE prospects SET statut=?, date_relance=? WHERE id=?',
            (nouveau_statut, nouvelle_relance, prospect_id)
        )
        conn.commit()
        conn.close()
        return redirect(url_for('voir_prospect', id=prospect_id))
    c.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,))
    prospect = c.fetchone()
    conn.close()
    return render_template('editer_intervention.html', intervention=intervention, prospect=prospect)

@app.route('/intervention/<int:id>/supprimer', methods=['POST'])
@login_required
def supprimer_intervention(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT prospect_id FROM interventions WHERE id = ?", (id,))
    res = c.fetchone()
    prospect_id = res['prospect_id'] if res else None
    c.execute("DELETE FROM intervention_attachments WHERE intervention_id = ?", (id,))
    c.execute("DELETE FROM interventions WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    if prospect_id:
        return redirect(url_for('voir_prospect', id=prospect_id))
    return redirect(url_for('index'))

# ─────────────────────────────────────────────
# ACTIONS MULTIPLES
# ─────────────────────────────────────────────

@app.route('/action_multiple', methods=['GET', 'POST'])
@login_required
def action_multiple():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if request.method == 'POST':
        prospect_ids = request.form.getlist('prospect_ids')
        date_intervention = request.form.get('date_intervention', '').strip()
        type_contact = request.form.get('type_contact', '').strip()
        compte_rendu = request.form.get('compte_rendu', '').strip()
        nouveau_statut = request.form.get('nouveau_statut', '').strip()
        date_relance = request.form.get('date_relance', '').strip()
        nb = 0
        uid = session['user_id']
        for pid in prospect_ids:
            if not pid.isdigit():
                continue
            # Vérifier propriété
            c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (pid, uid))
            if not c.fetchone():
                continue
            nb += 1
            if date_intervention or type_contact or compte_rendu:
                c.execute(
                    "INSERT INTO interventions (prospect_id, date_intervention, type_contact, compte_rendu, username) VALUES (?, ?, ?, ?, ?)",
                    (pid, date_intervention, type_contact, compte_rendu, session.get('username', ''))
                )
            if nouveau_statut or date_relance:
                c.execute("SELECT statut FROM prospects WHERE id = ?", (pid,))
                p = c.fetchone()
                if p:
                    statut_final = nouveau_statut if nouveau_statut else p['statut']
                    relance_finale = date_relance if date_relance else None
                    ancien_statut = p['statut']
                    sets = ['date_relance=?']
                    vals = [relance_finale]
                    if nouveau_statut:
                        sets.insert(0, 'statut=?')
                        vals.insert(0, statut_final)
                    vals.append(pid)
                    c.execute(f"UPDATE prospects SET {', '.join(sets)} WHERE id=?", vals)
                    if nouveau_statut and statut_final != ancien_statut:
                        c.execute(
                            "INSERT INTO historique_statut (prospect_id, ancien_statut, nouveau_statut, username) VALUES (?, ?, ?, ?)",
                            (pid, ancien_statut, statut_final, session.get('username'))
                        )
        conn.commit()
        conn.close()
        success = f"Action appliquée sur {nb} candidature(s) avec succès."
        conn2 = sqlite3.connect(DB_NAME)
        conn2.row_factory = sqlite3.Row
        c2 = conn2.cursor()
        c2.execute("SELECT id, etablissement, poste, categorie, statut, type_contrat FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=? ORDER BY etablissement ASC", (session['user_id'],))
        prospects = c2.fetchall()
        conn2.close()
        return render_template('action_multiple.html', prospects=prospects,
                               today=datetime.now().strftime('%Y-%m-%d'), success=success)
    c.execute("SELECT id, etablissement, poste, categorie, statut, type_contrat FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=? ORDER BY etablissement ASC", (session['user_id'],))
    prospects = c.fetchall()
    conn.close()
    return render_template('action_multiple.html', prospects=prospects,
                           today=datetime.now().strftime('%Y-%m-%d'))

# ─────────────────────────────────────────────
# CALENDRIER
# ─────────────────────────────────────────────

@app.route('/calendrier')
@login_required
def calendrier():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    uid = session['user_id']
    c.execute("SELECT id, etablissement, poste, date_relance FROM prospects WHERE date_relance IS NOT NULL AND date_relance != '' AND statut NOT IN ('Refus','Désistement','Offre reçue') AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    relances = c.fetchall()
    c.execute("SELECT * FROM evenements_perso WHERE user_id=? ORDER BY date_event ASC", (uid,))
    evenements = c.fetchall()
    c.execute("SELECT id, etablissement, poste FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=? ORDER BY etablissement ASC", (uid,))
    tous_prospects = c.fetchall()
    ical_token = c.execute("SELECT ical_token FROM users WHERE id=?", (uid,)).fetchone()
    ical_token = ical_token['ical_token'] if ical_token else None
    conn.close()
    return render_template('calendrier.html', relances=relances, evenements=evenements,
                           tous_prospects=tous_prospects, ical_token=ical_token)

@app.route('/calendrier/ajouter', methods=['POST'])
@login_required
def ajouter_event_calendrier():
    titre = request.form.get('titre', '').strip()
    date_event = request.form.get('date_event', '').strip()
    type_event = request.form.get('type_event', '').strip()
    if not titre or not date_event or not type_event:
        return redirect(url_for('calendrier'))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT INTO evenements_perso (titre, date_event, type_event, user_id) VALUES (?, ?, ?, ?)", (titre, date_event, type_event, session['user_id']))
    conn.commit()
    conn.close()
    return redirect(url_for('calendrier'))

@app.route('/calendrier/supprimer/<int:id>', methods=['POST'])
@login_required
def supprimer_event_calendrier(id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM evenements_perso WHERE id=? AND user_id=?", (id, session['user_id']))
    conn.commit()
    conn.close()
    return redirect(url_for('calendrier'))

@app.route('/calendrier/feed/<token>.ics')
def ical_feed(token):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    user = c.execute("SELECT id, username FROM users WHERE ical_token=? AND approved=1", (token,)).fetchone()
    if not user:
        conn.close()
        abort(404)
    uid = user['id']
    company_row = c.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
    company_name = company_row['value'] if company_row else 'OpenSuivi'

    cal = Calendar()
    cal.add('prodid', f'-//{company_name}//Suivi Emploi//FR')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    cal.add('x-wr-calname', vText(f'{company_name} — {user["username"]}'))
    cal.add('x-wr-timezone', vText('Europe/Paris'))
    cal.add('x-wr-caldesc', vText('Relances et événements de suivi emploi'))

    # Événements manuels
    events = c.execute(
        "SELECT id, titre, date_event, type_event FROM evenements_perso WHERE user_id=?", (uid,)
    ).fetchall()
    for e in events:
        try:
            dt = datetime.strptime(e['date_event'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        ev = Event()
        ev.add('uid', vText(f'event-{e["id"]}@opensuivi'))
        ev.add('summary', vText(f'{e["type_event"]} : {e["titre"]}'))
        ev.add('dtstart', dt)
        ev.add('dtend', dt)
        cal.add_component(ev)

    # Relances candidatures
    relances = c.execute(
        "SELECT id, etablissement, poste, date_relance FROM prospects "
        "WHERE date_relance IS NOT NULL AND date_relance != '' "
        "AND statut NOT IN ('Refus','Désistement','Offre reçue') "
        "AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,)
    ).fetchall()
    for r in relances:
        try:
            dt = datetime.strptime(r['date_relance'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        ev = Event()
        ev.add('uid', vText(f'relance-{r["id"]}@opensuivi'))
        label = f'{r["etablissement"]}'
        if r['poste']:
            label += f' — {r["poste"]}'
        ev.add('summary', vText(f'Relance : {label}'))
        ev.add('dtstart', dt)
        ev.add('dtend', dt)
        cal.add_component(ev)

    # Dates limites de candidature
    limites = c.execute(
        "SELECT id, etablissement, poste, date_limite_candidature FROM prospects "
        "WHERE date_limite_candidature IS NOT NULL AND date_limite_candidature != '' "
        "AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,)
    ).fetchall()
    for l in limites:
        try:
            dt = datetime.strptime(l['date_limite_candidature'], '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        ev = Event()
        ev.add('uid', vText(f'limite-{l["id"]}@opensuivi'))
        label = f'{l["etablissement"]}'
        if l['poste']:
            label += f' — {l["poste"]}'
        ev.add('summary', vText(f'Date limite : {label}'))
        ev.add('dtstart', dt)
        ev.add('dtend', dt)
        cal.add_component(ev)

    conn.close()
    return Response(
        cal.to_ical(),
        mimetype='text/calendar',
        headers={'Content-Disposition': 'inline; filename="opensuivi.ics"'}
    )

# ─────────────────────────────────────────────
# KANBAN
# ─────────────────────────────────────────────

@app.route('/kanban')
@login_required
def kanban():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=? ORDER BY date_relance ASC NULLS LAST, etablissement ASC", (session['user_id'],))
    prospects = c.fetchall()
    conn.close()
    return render_template('kanban.html', prospects=prospects, colonnes=STATUTS)

@app.route('/prospect/<int:id>/statut', methods=['POST'])
@login_required
def update_statut(id):
    data = request.get_json()
    nouveau_statut = (data or {}).get('statut', '').strip()
    if not nouveau_statut:
        return {'ok': False}, 400
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT statut FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    row = c.fetchone()
    if not row:
        conn.close()
        return {'ok': False}, 404
    ancien_statut = row['statut']
    c.execute("UPDATE prospects SET statut = ? WHERE id = ?", (nouveau_statut, id))
    if nouveau_statut != ancien_statut:
        c.execute("INSERT INTO historique_statut (prospect_id, ancien_statut, nouveau_statut, username) VALUES (?, ?, ?, ?)",
                  (id, ancien_statut, nouveau_statut, session.get('username')))
    conn.commit()
    conn.close()
    return {'ok': True}

# ─────────────────────────────────────────────
# CALCULATRICE
# ─────────────────────────────────────────────

@app.route('/calculatrice')
@login_required
def calculatrice():
    return render_template('calculatrice.html')

# ─────────────────────────────────────────────
# EXPORT / IMPORT CSV
# ─────────────────────────────────────────────

@app.route('/export')
@login_required
def export_csv():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE user_id=? ORDER BY date_ajout DESC", (session['user_id'],))
    prospects = c.fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['ID', 'Entreprise', 'Contact', 'Téléphone', 'Email', 'Ville', 'Secteur',
                     'Poste', 'Type contrat', 'Plateforme', 'Statut',
                     'CV envoyé', 'Lettre envoyée', 'Date relance', 'Commentaire', 'Date ajout'])
    for p in prospects:
        writer.writerow([
            p['id'], p['etablissement'], p['contact'], p['telephone'], p['email'],
            p['ville'], p['categorie'], p['poste'], p['type_contrat'], p['plateforme'],
            p['statut'],
            'Oui' if p['cv_envoye'] else 'Non',
            'Oui' if p['lettre_envoyee'] else 'Non',
            p['date_relance'], p['commentaire'], p['date_ajout']
        ])
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment; filename=candidatures_export.csv'})

@app.route('/export/xlsx')
@login_required
def export_xlsx():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM prospects WHERE user_id=? ORDER BY date_ajout DESC", (session['user_id'],))
    prospects = c.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatures"

    headers = ['ID', 'Entreprise', 'Poste', 'Ville', 'Type contrat', 'Plateforme', 'Secteur',
               'Statut', 'CV envoyé', 'LM envoyée', 'Date relance', 'Date limite',
               'Date début', 'Durée', 'Formation requise', 'Référence', 'Lien offre',
               'Contact', 'Email', 'Téléphone', 'Commentaire', 'Date ajout']

    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, p in enumerate(prospects, 2):
        ws.append([
            p['id'], p['etablissement'], p['poste'], p['ville'],
            p['type_contrat'], p['plateforme'], p['categorie'], p['statut'],
            'Oui' if p['cv_envoye'] else 'Non',
            'Oui' if p['lettre_envoyee'] else 'Non',
            p['date_relance'], p['date_limite_candidature'],
            p['date_debut'], p['duree_contrat'], p['formation_requise'],
            p['reference_offre'], p['lien_offre'],
            p['contact'], p['email'], p['telephone'],
            p['commentaire'], p['date_ajout']
        ])
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=candidatures_export.xlsx'})

@app.route('/stats')
@login_required
def stats():
    uid = session['user_id']
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    total = c.execute("SELECT COUNT(*) FROM prospects WHERE user_id=? AND (archived=0 OR archived IS NULL)", (uid,)).fetchone()[0]
    by_statut = c.execute("SELECT statut, COUNT(*) as n FROM prospects WHERE user_id=? AND (archived=0 OR archived IS NULL) GROUP BY statut ORDER BY n DESC", (uid,)).fetchall()
    by_plateforme = c.execute("SELECT plateforme, COUNT(*) as n FROM prospects WHERE user_id=? AND (archived=0 OR archived IS NULL) AND plateforme IS NOT NULL AND plateforme != '' GROUP BY plateforme ORDER BY n DESC", (uid,)).fetchall()
    by_secteur = c.execute("SELECT categorie, COUNT(*) as n FROM prospects WHERE user_id=? AND (archived=0 OR archived IS NULL) AND categorie IS NOT NULL AND categorie != '' GROUP BY categorie ORDER BY n DESC", (uid,)).fetchall()
    by_contrat = c.execute("SELECT type_contrat, COUNT(*) as n FROM prospects WHERE user_id=? AND (archived=0 OR archived IS NULL) AND type_contrat IS NOT NULL AND type_contrat != '' GROUP BY type_contrat ORDER BY n DESC", (uid,)).fetchall()

    entretiens = c.execute("SELECT COUNT(*) FROM prospects WHERE user_id=? AND statut IN ('Entretien téléphonique','Entretien présentiel','Test technique','Offre reçue') AND (archived=0 OR archived IS NULL)", (uid,)).fetchone()[0]
    offres = c.execute("SELECT COUNT(*) FROM prospects WHERE user_id=? AND statut='Offre reçue' AND (archived=0 OR archived IS NULL)", (uid,)).fetchone()[0]
    refus = c.execute("SELECT COUNT(*) FROM prospects WHERE user_id=? AND statut='Refus' AND (archived=0 OR archived IS NULL)", (uid,)).fetchone()[0]

    # Candidatures par semaine (12 dernières semaines)
    par_semaine = c.execute(
        "SELECT strftime('%Y-W%W', date_ajout) as semaine, COUNT(*) as n "
        "FROM prospects WHERE user_id=? AND date_ajout >= date('now', '-84 days') "
        "GROUP BY semaine ORDER BY semaine ASC", (uid,)
    ).fetchall()

    # Délai moyen entre envoi et premier entretien
    delai_row = c.execute(
        "SELECT AVG(julianday(h.horodatage) - julianday(p.date_ajout)) "
        "FROM historique_statut h JOIN prospects p ON p.id=h.prospect_id "
        "WHERE p.user_id=? AND h.nouveau_statut IN ('Entretien téléphonique','Entretien présentiel')", (uid,)
    ).fetchone()
    delai_moyen = round(delai_row[0]) if delai_row and delai_row[0] else None

    conn.close()
    taux_reponse = round(entretiens * 100 / total) if total > 0 else 0
    return render_template('stats.html',
        total=total, by_statut=by_statut, by_plateforme=by_plateforme,
        by_secteur=by_secteur, by_contrat=by_contrat,
        entretiens=entretiens, offres=offres, refus=refus,
        taux_reponse=taux_reponse, par_semaine=par_semaine,
        delai_moyen=delai_moyen)

def _csv_find(row, *aliases):
    row_lower = {k.lower().strip(): v for k, v in row.items()}
    for alias in aliases:
        v = row_lower.get(alias.lower().strip())
        if v is not None:
            return v.strip() if isinstance(v, str) else v
    return ''

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_csv():
    error = None
    success = None
    detected_cols = []
    skipped = 0
    if request.method == 'POST':
        f = request.files.get('fichier')
        sep = request.form.get('separateur', ';')
        if not f or not f.filename.lower().endswith('.csv'):
            error = "Veuillez sélectionner un fichier CSV."
        else:
            try:
                content = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content), delimiter=sep)
                detected_cols = list(reader.fieldnames or [])
                count = 0
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                for row in reader:
                    etab = _csv_find(row, 'Entreprise', 'etablissement', 'Nom', 'Société', 'Company', 'Établissement')
                    if not etab:
                        skipped += 1
                        continue
                    cv_val = _csv_find(row, 'CV envoyé', 'cv_envoye', 'CV').lower()
                    lettre_val = _csv_find(row, 'Lettre envoyée', 'lettre_envoyee', 'Lettre').lower()
                    date_rel = _csv_find(row, 'Date relance', 'date_relance', 'Relance') or None
                    c.execute(
                        """INSERT INTO prospects
                           (etablissement, contact, telephone, email, ville, code_postal,
                            categorie, poste, type_contrat, plateforme, statut,
                            date_relance, commentaire, cv_envoye, lettre_envoyee, user_id)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (etab,
                         _csv_find(row, 'Contact', 'Nom du contact', 'contact'),
                         _csv_find(row, 'Téléphone', 'Telephone', 'Tel', 'telephone'),
                         _csv_find(row, 'Email', 'email', 'Mail'),
                         _csv_find(row, 'Ville', 'ville', 'City'),
                         _csv_find(row, 'Code Postal', 'code_postal', 'CP'),
                         _csv_find(row, 'Secteur', 'categorie', 'Catégorie', 'Category'),
                         _csv_find(row, 'Poste', 'poste', 'Intitulé', 'Job'),
                         _csv_find(row, 'Type contrat', 'type_contrat', 'Contrat'),
                         _csv_find(row, 'Plateforme', 'plateforme', 'Source'),
                         _csv_find(row, 'Statut', 'statut', 'Status'),
                         date_rel,
                         _csv_find(row, 'Commentaire', 'commentaire', 'Notes'),
                         1 if cv_val in ('oui', '1', 'true') else 0,
                         1 if lettre_val in ('oui', '1', 'true') else 0,
                         session['user_id'],
                        )
                    )
                    count += 1
                conn.commit()
                conn.close()
                msg = f"{count} candidature(s) importée(s) avec succès."
                if skipped:
                    msg += f" {skipped} ligne(s) ignorée(s) (sans entreprise)."
                success = msg
            except Exception as e:
                error = f"Erreur lors de l'import : {e}"
    return render_template('import.html', error=error, success=success,
                           detected_cols=detected_cols, skipped=skipped)

# ─────────────────────────────────────────────
# NOTE RAPIDE
# ─────────────────────────────────────────────

@app.route('/prospect/<int:id>/note-rapide', methods=['POST'])
@login_required
def note_rapide(id):
    note = request.form.get('note', '').strip()
    if note:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("INSERT INTO interventions (prospect_id, date_intervention, type_contact, compte_rendu, username) VALUES (?, ?, 'Note', ?, ?)",
                  (id, datetime.now().strftime('%Y-%m-%d'), note, session.get('username', '')))
        conn.commit()
        conn.close()
    return redirect(request.referrer or url_for('base_donnees'))

# ─────────────────────────────────────────────
# DOCUMENTS CV / LM PAR CANDIDATURE
# ─────────────────────────────────────────────

@app.route('/prospect/<int:id>/upload-doc', methods=['POST'])
@login_required
def upload_prospect_doc(id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    if not c.fetchone():
        conn.close()
        abort(403)
    doc_type = request.form.get('doc_type', '').strip()
    if doc_type not in ('cv', 'lm'):
        conn.close()
        return redirect(url_for('voir_prospect', id=id))
    if 'file' not in request.files:
        conn.close()
        return redirect(url_for('voir_prospect', id=id))
    f = request.files['file']
    if not f or not f.filename:
        conn.close()
        return redirect(url_for('voir_prospect', id=id))
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in {'pdf', 'doc', 'docx'}:
        conn.close()
        return redirect(url_for('voir_prospect', id=id) + '?doc_error=1')
    # Supprimer l'ancien fichier s'il existe
    old = c.execute("SELECT filename FROM prospect_documents WHERE prospect_id=? AND doc_type=?", (id, doc_type)).fetchone()
    if old:
        old_path = os.path.join(UPLOAD_FOLDER, old['filename'])
        if os.path.exists(old_path):
            os.remove(old_path)
        c.execute("DELETE FROM prospect_documents WHERE prospect_id=? AND doc_type=?", (id, doc_type))
    unique_name = f"{secrets.token_hex(12)}.{ext}"
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    f.save(os.path.join(UPLOAD_FOLDER, unique_name))
    c.execute("INSERT INTO prospect_documents (prospect_id, doc_type, filename, original_name) VALUES (?,?,?,?)",
              (id, doc_type, unique_name, f.filename))
    conn.commit()
    conn.close()
    return redirect(url_for('voir_prospect', id=id) + '#documents')

@app.route('/prospect/<int:id>/doc/<doc_type>')
@login_required
def download_prospect_doc(id, doc_type):
    if doc_type not in ('cv', 'lm'):
        abort(404)
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    if not c.fetchone():
        conn.close()
        abort(403)
    doc = c.execute("SELECT * FROM prospect_documents WHERE prospect_id=? AND doc_type=?", (id, doc_type)).fetchone()
    conn.close()
    if not doc:
        abort(404)
    return send_from_directory(UPLOAD_FOLDER, doc['filename'],
                               as_attachment=True, download_name=doc['original_name'])

@app.route('/prospect/<int:id>/doc/<doc_type>/supprimer', methods=['POST'])
@login_required
def supprimer_prospect_doc(id, doc_type):
    if doc_type not in ('cv', 'lm'):
        abort(404)
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    if not c.fetchone():
        conn.close()
        abort(403)
    doc = c.execute("SELECT * FROM prospect_documents WHERE prospect_id=? AND doc_type=?", (id, doc_type)).fetchone()
    if doc:
        fpath = os.path.join(UPLOAD_FOLDER, doc['filename'])
        if os.path.exists(fpath):
            os.remove(fpath)
        c.execute("DELETE FROM prospect_documents WHERE id=?", (doc['id'],))
        conn.commit()
    conn.close()
    return redirect(url_for('voir_prospect', id=id) + '#documents')

# ─────────────────────────────────────────────
# NOTES D'ENTRETIEN
# ─────────────────────────────────────────────

@app.route('/prospect/<int:id>/notes-entretien', methods=['POST'])
@login_required
def sauvegarder_notes_entretien(id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id FROM prospects WHERE id=? AND user_id=?", (id, session['user_id']))
    if not c.fetchone():
        conn.close()
        abort(403)
    notes = request.form.get('notes_entretien', '').strip()
    c.execute("UPDATE prospects SET notes_entretien=? WHERE id=?", (notes, id))
    conn.commit()
    conn.close()
    return redirect(url_for('voir_prospect', id=id) + '#notes-entretien')

# ─────────────────────────────────────────────
# API REST
# ─────────────────────────────────────────────

@app.route('/api/prospects')
@login_required
def api_prospects():
    q = request.args.get('q', '').strip()
    statut = request.args.get('statut', '').strip()
    secteur = request.args.get('secteur', '').strip()
    try:
        page = max(1, int(request.args.get('page', 1) or 1))
        per_page = min(500, max(1, int(request.args.get('per_page', 100) or 100)))
    except (ValueError, TypeError):
        page, per_page = 1, 100
    conditions = ["(archived IS NULL OR archived = 0)", "user_id = ?"]
    params = [session['user_id']]
    if q:
        conditions.append("(etablissement LIKE ? OR contact LIKE ? OR poste LIKE ? OR email LIKE ?)")
        like = f'%{q}%'; params += [like, like, like, like]
    if statut:
        conditions.append("statut = ?"); params.append(statut)
    if secteur:
        conditions.append("categorie = ?"); params.append(secteur)
    where = "WHERE " + " AND ".join(conditions)
    offset = (page - 1) * per_page
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(f"SELECT COUNT(*) FROM prospects {where}", params)
    total = c.fetchone()[0]
    c.execute(f"""SELECT id, etablissement, contact, email, telephone, ville,
                         categorie, poste, type_contrat, plateforme, statut,
                         cv_envoye, lettre_envoyee, date_relance
                  FROM prospects {where} ORDER BY etablissement ASC LIMIT ? OFFSET ?""",
              params + [per_page, offset])
    data = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(data=data, total=total, page=page, per_page=per_page,
                   pages=max(1, (total + per_page - 1) // per_page))

@app.route('/api/stats')
@login_required
def api_stats():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    uid = session['user_id']
    c.execute("SELECT COUNT(*) as total FROM prospects WHERE (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    total = c.fetchone()['total']
    c.execute("SELECT COUNT(*) as cv FROM prospects WHERE cv_envoye=1 AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    cv = c.fetchone()['cv']
    c.execute("SELECT COUNT(*) as entretiens FROM prospects WHERE statut IN ('Entretien téléphonique','Entretien présentiel','Test technique') AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    entretiens = c.fetchone()['entretiens']
    c.execute("SELECT COUNT(*) as offres FROM prospects WHERE statut='Offre reçue' AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    offres = c.fetchone()['offres']
    c.execute("SELECT COUNT(*) as relances FROM prospects WHERE date_relance <= date('now') AND statut NOT IN ('Offre reçue','Refus','Désistement') AND (archived=0 OR archived IS NULL) AND user_id=?", (uid,))
    relances = c.fetchone()['relances']
    conn.close()
    return jsonify(total=total, cv_envoyes=cv, entretiens=entretiens, offres=offres, relances_en_retard=relances)

@app.route('/api/check-duplicate')
@login_required
def api_check_duplicate():
    nom   = request.args.get('nom', '').strip()
    email = request.args.get('email', '').strip()
    pid   = request.args.get('exclude_id', '')
    conditions, params = [], []
    if nom:
        conditions.append("etablissement LIKE ?"); params.append(f'%{nom}%')
    if email:
        conditions.append("email = ?"); params.append(email)
    if not conditions:
        return jsonify(doublons=[])
    where = "WHERE (" + " OR ".join(conditions) + ") AND (archived=0 OR archived IS NULL) AND user_id=?"
    params.append(session['user_id'])
    if pid and pid.isdigit():
        where += " AND id != ?"; params.append(pid)
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(f"SELECT id, etablissement, poste, email, statut FROM prospects {where} LIMIT 5", params)
    doublons = [dict(r) for r in c.fetchall()]
    conn.close()
    return jsonify(doublons=doublons)

# ─────────────────────────────────────────────
# EXTRACTION AUTOMATIQUE D'OFFRE (OpenAI)
# ─────────────────────────────────────────────

SECTEURS_LIST = ['Informatique/Tech', 'Commerce/Vente', 'Finance/Comptabilité',
                 'Marketing/Communication', 'RH/Formation', 'Santé',
                 'Industrie/Production', 'Bâtiment/Travaux', 'Transport/Logistique', 'Autre']
TYPES_CONTRAT_LIST = ['Alternance', 'Stage', 'CDI', 'CDD', 'Freelance']
PLATEFORMES_LIST = ['LinkedIn', 'Indeed', 'France Travail', 'Direct', 'Autre']
MODES_CAND_LIST = ['Via plateforme', 'Via email', 'Via courrier', 'Direct / Spontanée']

# Sélecteurs spécifiques d'abord, génériques en dernier
_CONTENT_SELECTORS = [
    '#colD', '#col-right', '#col_right', '#rightcol',
    '#job-detail', '#jobDetail', '#offer-detail', '#offerDetail',
    '.job-detail', '.offer-detail', '.job-description', '.jobDescription',
    '.offre-detail', '.ts-offer', '[class*="job-content"]', '[class*="offer-content"]',
    '[id*="offer-detail"]', '[id*="job-detail"]',
    'article', 'main',
]

def _scrape_workday(url):
    """Workday est une SPA : on appelle directement leur API JSON interne."""
    from urllib.parse import urlparse
    parsed = urlparse(url)
    hostname = parsed.hostname  # ex: ratp.wd3.myworkdayjobs.com
    tenant = hostname.split('.')[0]  # ex: ratp

    # Chemin : /fr-FR/RATP_Externe/job/... → supprimer le préfixe locale
    path_parts = parsed.path.lstrip('/').split('/')
    # Le premier segment est la locale (ex: fr-FR) si court et contient '-'
    if path_parts and '-' in path_parts[0] and len(path_parts[0]) <= 6:
        path_parts = path_parts[1:]

    api_path = '/wday/cxs/' + tenant + '/' + '/'.join(path_parts)
    api_url = f"https://{hostname}{api_path}"

    resp = http_requests.get(api_url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'application/json',
    }, timeout=12)
    resp.raise_for_status()
    info = resp.json().get('jobPostingInfo', {})

    desc_html = info.get('jobDescription', '')
    desc_text = BeautifulSoup(desc_html, 'html.parser').get_text(separator='\n', strip=True)

    parts = []
    if info.get('title'):       parts.append(f"Poste : {info['title']}")
    if info.get('location'):    parts.append(f"Lieu : {info['location']}")
    if info.get('startDate'):   parts.append(f"Date de début : {info['startDate']}")
    if info.get('jobReqId'):    parts.append(f"Référence : {info['jobReqId']}")
    if info.get('remoteType'):  parts.append(f"Télétravail : {info['remoteType']}")
    if desc_text:               parts.append(f"\nDescription :\n{desc_text}")

    return '\n'.join(parts)


def _scrape_url(url):
    if 'myworkdayjobs.com' in url:
        return _scrape_workday(url)

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept-Language': 'fr-FR,fr;q=0.9',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }
    resp = http_requests.get(url, headers=headers, timeout=12, allow_redirects=True)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, 'html.parser')

    # Supprimer le bruit non-textuel et les listes déroulantes
    for tag in soup(['script', 'style', 'noscript', 'iframe', 'select', 'option', 'datalist']):
        tag.decompose()

    # Chercher le conteneur principal, valider qu'il a du contenu substantiel
    content_el = None
    for sel in _CONTENT_SELECTORS:
        el = soup.select_one(sel)
        if el:
            txt = el.get_text(strip=True)
            if len(txt) > 300:
                content_el = el
                break

    target = content_el if content_el else soup.body or soup

    # Supprimer les éléments de navigation/décoration dans le conteneur retenu
    for tag in target(['nav', 'header', 'footer', 'aside']):
        tag.decompose()

    text = target.get_text(separator='\n', strip=True)
    # Filtrer les lignes trop courtes (boutons, libellés isolés) et les doublons consécutifs
    lines, prev = [], ''
    for l in text.splitlines():
        l = l.strip()
        if len(l) > 2 and l != prev:
            lines.append(l)
            prev = l

    limit = 600 if content_el else 450
    return '\n'.join(lines[:limit])

def _extract_with_openai(text_content, source_url=''):
    api_key = get_openai_key()
    if not api_key:
        return None, "Clé API OpenAI non configurée (Admin → Personnalisation → Intégrations)"
    client = OpenAI(api_key=api_key)

    url_hint = f"\nURL source : {source_url}" if source_url else ""

    prompt = f"""Tu es un assistant expert en extraction d'informations depuis des offres d'emploi françaises.{url_hint}

Voici le contenu brut de l'offre :
---
{text_content[:12000]}
---

Extrais TOUTES les informations disponibles en JSON. Laisse "" si une info est absente.

Champs à extraire :
- entreprise : nom complet de l'organisme/entreprise qui recrute
- poste : intitulé exact du poste (tel qu'écrit dans l'annonce, sans H/F)
- ville : ville principale du lieu de travail.
  RÈGLES IMPORTANTES pour la ville :
  * Si tu vois "LPO/LP/LEG/LEGT/LPO XXX - VILLE" ou "LYCEE XXX - VILLE", extrais la VILLE après le tiret
  * Si tu vois "Résidence administrative : VILLE", utilise cette ville
  * Si tu vois "Maison de la Région de VILLE", utilise cette ville
  * Priorité : résidence administrative > nom du lycée/site > maison de région
- code_postal : code postal si mentionné
- adresse : lieu de travail précis.
  RÈGLES : si un nom de lycée/établissement/site est mentionné (ex: "LPO STEPHANE HESSEL - EPERNAY" ou "Maison de la Région de Châlons"), mets-le ici comme adresse.
  Format : "Nom de l'établissement, Ville"
- type_contrat : parmi {TYPES_CONTRAT_LIST} (Alternance si apprentissage/contrat d'apprentissage)
- secteur : parmi {SECTEURS_LIST} (le plus pertinent)
- plateforme : parmi {PLATEFORMES_LIST} (déduit depuis l'URL ou le contenu)
- mode_candidature : parmi {MODES_CAND_LIST} (Via plateforme si espace candidat en ligne, Via email si email fourni)
- contact : prénom et nom du recruteur/référent RH si mentionné
- email_contact : adresse email de contact si mentionnée
- telephone : numéro de téléphone si mentionné
- reference_offre : numéro/référence de l'offre (ex: 2026-5767)
- date_limite : date limite de candidature au format YYYY-MM-DD (ex: 2026-05-17), "" si absente
- date_debut : date de début du poste au format YYYY-MM-DD, "" si absente
- duree_contrat : durée du contrat (ex: "2 ans", "6 mois"), "" si non précisée
- formation_requise : niveau ou diplôme requis (ex: "BAC+2 GMSI ou TSR", "Bac+3"), "" si absent
- commentaire : résumé structuré de l'offre : missions principales + profil recherché (3-5 lignes max)

Réponds UNIQUEMENT avec le JSON valide, sans texte autour, sans markdown.
Exemples de dates : "17/05/2026" → "2026-05-17", "1er septembre 2026" → "2026-09-01"."""

    response = client.chat.completions.create(
        model='gpt-4o-mini',
        messages=[{'role': 'user', 'content': prompt}],
        temperature=0.1,
        max_tokens=1200,
    )
    raw = response.choices[0].message.content.strip()
    raw = re.sub(r'^```json\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'^```\s*', '', raw, flags=re.MULTILINE)
    raw = re.sub(r'\s*```$', '', raw, flags=re.MULTILINE)
    data = json.loads(raw)
    return data, None

@app.route('/api/extract-offre', methods=['POST'])
@login_required
def api_extract_offre():
    body = request.get_json(silent=True) or {}
    source = (body.get('source') or '').strip()
    if not source:
        return jsonify(error="Source vide"), 400

    is_url = source.startswith('http://') or source.startswith('https://')

    if is_url:
        try:
            text_content = _scrape_url(source)
        except Exception as e:
            return jsonify(error=f"Impossible de lire la page : {str(e)}. Essayez de coller le texte de l'offre directement."), 400
        source_url = source
    else:
        text_content = source
        source_url = ''

    if len(text_content.strip()) < 30:
        return jsonify(error="Contenu trop court pour être analysé."), 400

    data, err = _extract_with_openai(text_content, source_url)
    if err:
        return jsonify(error=err), 400

    # Toujours retourner l'URL si c'était une URL
    if is_url:
        data['lien_offre'] = source

    return jsonify(data)

# ─────────────────────────────────────────────
# IA — OUTILS PAR CANDIDATURE
# ─────────────────────────────────────────────

def _get_prospect_for_ia(prospect_id):
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    p = c.execute("SELECT * FROM prospects WHERE id=? AND user_id=?",
                  (prospect_id, session['user_id'])).fetchone()
    conn.close()
    return p

@app.route('/api/generer-lm/<int:id>', methods=['POST'])
@login_required
def api_generer_lm(id):
    key = get_openai_key()
    if not key:
        return jsonify(error="Clé OpenAI non configurée."), 400
    p = _get_prospect_for_ia(id)
    if not p:
        abort(403)
    uid = session['user_id']
    profil = {
        'type_contrat': get_user_setting(uid, 'profil_type_contrat', ''),
        'niveau':       get_user_setting(uid, 'profil_niveau', ''),
        'specialite':   get_user_setting(uid, 'profil_specialite', ''),
        'ecole':        get_user_setting(uid, 'profil_ecole', ''),
        'rythme':       get_user_setting(uid, 'profil_rythme', ''),
        'dispo':        get_user_setting(uid, 'profil_dispo', ''),
        'localisation': get_user_setting(uid, 'profil_localisation', ''),
        'salaire':      get_user_setting(uid, 'profil_salaire', ''),
    }
    prompt = f"""Tu es un assistant expert en recrutement français. Rédige une lettre de motivation professionnelle et personnalisée en français pour cette candidature.

Poste : {p['poste'] or 'Non précisé'}
Entreprise : {p['etablissement']}
Ville : {p['ville'] or 'Non précisée'}
Type de contrat : {p['type_contrat'] or 'Non précisé'}
Secteur : {p['categorie'] or 'Non précisé'}
Formation requise : {p['formation_requise'] or 'Non précisée'}
Description du poste : {(p['commentaire'] or '')[:800]}

Profil du candidat :
- Niveau : {profil['niveau'] or 'Non précisé'}
- Spécialité : {profil['specialite'] or 'Non précisée'}
- École/Formation : {profil['ecole'] or 'Non précisée'}
- Type de contrat recherché : {profil['type_contrat'] or 'Non précisé'}
- Disponibilité : {profil['dispo'] or 'Non précisée'}
- Localisation : {profil['localisation'] or 'Non précisée'}

Rédige une lettre de motivation structurée (introduction, développement, conclusion) d'environ 300 mots. Sois professionnel, authentique et précis. Ne mets pas d'en-tête ni de formule de politesse finale elaborate, juste le corps de la lettre."""
    try:
        client = OpenAI(api_key=key)
        resp = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
            max_tokens=900,
            temperature=0.7,
        )
        lm = resp.choices[0].message.content.strip()
        return jsonify(lm=lm)
    except Exception as e:
        return jsonify(error=str(e)), 400

@app.route('/api/questions-entretien/<int:id>', methods=['POST'])
@login_required
def api_questions_entretien(id):
    key = get_openai_key()
    if not key:
        return jsonify(error="Clé OpenAI non configurée."), 400
    p = _get_prospect_for_ia(id)
    if not p:
        abort(403)
    prompt = f"""Tu es un expert en recrutement français. Pour cette offre d'emploi, génère les 10 questions d'entretien les plus probables que le recruteur va poser, avec pour chacune une courte piste de réponse.

Poste : {p['poste'] or 'Non précisé'}
Entreprise : {p['etablissement']}
Secteur : {p['categorie'] or 'Non précisé'}
Type de contrat : {p['type_contrat'] or 'Non précisé'}
Formation requise : {p['formation_requise'] or 'Non précisée'}
Description : {(p['commentaire'] or '')[:600]}

Format de réponse : liste numérotée, chaque item = "Question ?" suivi d'une piste de réponse courte (1-2 phrases) en italique."""
    try:
        client = OpenAI(api_key=key)
        resp = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
            max_tokens=1000,
            temperature=0.6,
        )
        questions = resp.choices[0].message.content.strip()
        return jsonify(questions=questions)
    except Exception as e:
        return jsonify(error=str(e)), 400

@app.route('/api/score-compat/<int:id>', methods=['POST'])
@login_required
def api_score_compat(id):
    key = get_openai_key()
    if not key:
        return jsonify(error="Clé OpenAI non configurée."), 400
    p = _get_prospect_for_ia(id)
    if not p:
        abort(403)
    uid = session['user_id']
    profil = {
        'type_contrat': get_user_setting(uid, 'profil_type_contrat', ''),
        'niveau':       get_user_setting(uid, 'profil_niveau', ''),
        'specialite':   get_user_setting(uid, 'profil_specialite', ''),
        'ecole':        get_user_setting(uid, 'profil_ecole', ''),
        'rythme':       get_user_setting(uid, 'profil_rythme', ''),
        'dispo':        get_user_setting(uid, 'profil_dispo', ''),
        'localisation': get_user_setting(uid, 'profil_localisation', ''),
        'salaire':      get_user_setting(uid, 'profil_salaire', ''),
    }
    prompt = f"""Tu es un expert en recrutement français. Analyse la compatibilité entre ce profil de candidat et cette offre d'emploi.

OFFRE :
Poste : {p['poste'] or 'Non précisé'}
Entreprise : {p['etablissement']}
Type contrat : {p['type_contrat'] or 'Non précisé'}
Secteur : {p['categorie'] or 'Non précisé'}
Formation requise : {p['formation_requise'] or 'Non précisée'}
Description : {(p['commentaire'] or '')[:600]}

PROFIL CANDIDAT :
Niveau : {profil['niveau'] or 'Non précisé'}
Spécialité : {profil['specialite'] or 'Non précisée'}
École : {profil['ecole'] or 'Non précisée'}
Type contrat recherché : {profil['type_contrat'] or 'Non précisé'}
Rythme : {profil['rythme'] or 'Non précisé'}
Disponibilité : {profil['dispo'] or 'Non précisée'}
Localisation : {profil['localisation'] or 'Non précisée'}

Réponds en JSON avec exactement ce format :
{{"score": <entier 0-100>, "points_forts": ["...", "..."], "points_faibles": ["...", "..."], "conseil": "..."}}"""
    try:
        client = OpenAI(api_key=key)
        resp = client.chat.completions.create(
            model='gpt-4o-mini',
            messages=[{'role': 'user', 'content': prompt}],
            max_tokens=500,
            temperature=0.4,
            response_format={"type": "json_object"},
        )
        import json as _json
        data = _json.loads(resp.choices[0].message.content)
        return jsonify(data)
    except Exception as e:
        return jsonify(error=str(e)), 400

# ─────────────────────────────────────────────
# PROFIL DE RECHERCHE
# ─────────────────────────────────────────────

@app.route('/profil-recherche', methods=['GET', 'POST'])
@login_required
def profil_recherche():
    uid = session['user_id']
    success = None
    if request.method == 'POST':
        fields = ['profil_type_contrat', 'profil_niveau', 'profil_specialite',
                  'profil_ecole', 'profil_rythme', 'profil_dispo',
                  'profil_localisation', 'profil_salaire']
        for field in fields:
            set_user_setting(uid, field, request.form.get(field, '').strip())
        success = "Profil de recherche enregistré."
    profil = {
        'type_contrat': get_user_setting(uid, 'profil_type_contrat', ''),
        'niveau':       get_user_setting(uid, 'profil_niveau', ''),
        'specialite':   get_user_setting(uid, 'profil_specialite', ''),
        'ecole':        get_user_setting(uid, 'profil_ecole', ''),
        'rythme':       get_user_setting(uid, 'profil_rythme', ''),
        'dispo':        get_user_setting(uid, 'profil_dispo', ''),
        'localisation': get_user_setting(uid, 'profil_localisation', ''),
        'salaire':      get_user_setting(uid, 'profil_salaire', ''),
    }
    return render_template('profil_recherche.html', profil=profil, success=success)

# ─────────────────────────────────────────────
# PROFIL UTILISATEUR
# ─────────────────────────────────────────────

@app.route('/profil', methods=['GET', 'POST'])
@login_required
def profil():
    error = None
    success = None
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    if request.method == 'POST':
        action = request.form.get('action', 'mdp')
        if action == 'signature':
            signature = request.form.get('signature', '')
            c.execute("UPDATE users SET signature = ? WHERE id = ?", (signature, session['user_id']))
            conn.commit()
            success = "Signature enregistrée."
        else:
            ancien_mdp = request.form.get('ancien_mdp', '')
            nouveau_mdp = request.form.get('nouveau_mdp', '')
            confirmation_mdp = request.form.get('confirmation_mdp', '')
            if not ancien_mdp or not nouveau_mdp or not confirmation_mdp:
                error = "Veuillez remplir tous les champs."
            elif nouveau_mdp != confirmation_mdp:
                error = "Les nouveaux mots de passe ne correspondent pas."
            elif len(nouveau_mdp) < 6:
                error = "Le nouveau mot de passe doit contenir au moins 6 caractères."
            else:
                c.execute("SELECT * FROM users WHERE id = ?", (session['user_id'],))
                user = c.fetchone()
                if user and check_password_hash(user['password'], ancien_mdp):
                    hashed_pw = generate_password_hash(nouveau_mdp)
                    c.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_pw, session['user_id']))
                    conn.commit()
                    success = "Mot de passe modifié avec succès."
                else:
                    error = "Ancien mot de passe incorrect."
    c.execute("SELECT signature FROM users WHERE id = ?", (session['user_id'],))
    row = c.fetchone()
    signature = row['signature'] if row else ''
    conn.close()
    return render_template('profil.html', error=error, success=success, current_signature=signature)

# ─────────────────────────────────────────────
# ADMINISTRATION
# ─────────────────────────────────────────────

def get_admin_users():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id, username, role, approved FROM users ORDER BY approved DESC, id ASC")
    users = c.fetchall()
    conn.close()
    return users

@app.route('/admin')
@login_required
@admin_required
def admin():
    tab = request.args.get('tab', 'users')
    success = request.args.get('success', '')
    error = request.args.get('error', '')
    company_info = {
        'name':      get_setting('company_name', ''),
        'forme':     get_setting('company_forme', ''),
        'email':     get_setting('company_email', ''),
        'telephone': get_setting('company_telephone', ''),
        'ville':     get_setting('company_ville', ''),
        'site':      get_setting('company_site', ''),
    }
    favicon_filename = get_setting('favicon_filename', '')
    favicon_url = f'/static/{favicon_filename}' if favicon_filename else ''
    raw_key = get_openai_key()
    openai_key_hint = f"{raw_key[:8]}...{raw_key[-4:]}" if len(raw_key) > 12 else ('(configurée)' if raw_key else '')
    return render_template('admin.html', users=get_admin_users(),
                           tab=tab, success=success, error=error,
                           company_info=company_info,
                           favicon_url=favicon_url,
                           openai_key_hint=openai_key_hint,
                           openai_key_set=bool(raw_key))

@app.route('/admin/settings', methods=['POST'])
@login_required
@admin_required
def admin_settings():
    fields = ['company_name', 'company_forme', 'company_email', 'company_telephone', 'company_ville', 'company_site']
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    for field in fields:
        val = request.form.get(field, '').strip()
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (field, val))
    # Clé OpenAI : ne remplace que si une valeur est fournie (évite d'écraser par vide)
    openai_key = request.form.get('openai_api_key', '').strip()
    if openai_key:
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('openai_api_key', ?)", (openai_key,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=perso&success=Paramètres+enregistrés')

@app.route('/admin/colors', methods=['POST'])
@login_required
@admin_required
def admin_colors():
    color_primary = request.form.get('color_primary', '').strip()
    color_accent  = request.form.get('color_accent', '').strip()
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    if color_primary:
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('color_primary', ?)", (color_primary,))
    if color_accent:
        c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('color_accent', ?)", (color_accent,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=perso&success=Couleurs+enregistrées')

@app.route('/admin/colors/reset')
@login_required
@admin_required
def admin_colors_reset():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM settings WHERE key IN ('color_primary', 'color_accent')")
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=perso&success=Couleurs+réinitialisées')

@app.route('/admin/upload-logo', methods=['POST'])
@login_required
@admin_required
def admin_upload_logo():
    if 'logo' not in request.files or request.files['logo'].filename == '':
        return redirect(url_for('admin') + '?tab=perso&error=Aucun+fichier+sélectionné')
    file = request.files['logo']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in {'png', 'jpg', 'jpeg', 'svg', 'ico', 'gif', 'webp'}:
        return redirect(url_for('admin') + '?tab=perso&error=Format+non+supporté')
    filename = f'logo_custom.{ext}'
    file.save(os.path.join(app.static_folder, filename))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('logo_filename', ?)", (filename,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=perso&success=Logo+enregistré')

@app.route('/admin/ajouter', methods=['POST'])
@login_required
@admin_required
def admin_ajouter_utilisateur():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'user')
    if not username or not password:
        return redirect(url_for('admin') + '?tab=users&error=Nom+et+mot+de+passe+obligatoires')
    if len(password) < 6:
        return redirect(url_for('admin') + '?tab=users&error=Mot+de+passe+trop+court+(6+min)')
    if role not in ('admin', 'commercial'):
        role = 'commercial'
    hashed_pw = generate_password_hash(password)
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    try:
        c.execute("INSERT INTO users (username, password, role, approved) VALUES (?, ?, ?, 1)", (username, hashed_pw, role))
        conn.commit()
        conn.close()
        return redirect(url_for('admin') + f'?tab=users&success=Utilisateur+{username}+créé')
    except sqlite3.IntegrityError:
        conn.close()
        return redirect(url_for('admin') + f'?tab=users&error=Nom+{username}+déjà+utilisé')

@app.route('/admin/approuver/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_approuver_utilisateur(id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("UPDATE users SET approved=1 WHERE id=?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=users&success=Compte+approuvé')

@app.route('/admin/refuser/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_refuser_utilisateur(id):
    if id == session.get('user_id'):
        return redirect(url_for('admin') + '?tab=users&error=Impossible+de+refuser+votre+propre+compte')
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id=? AND approved=0", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=users&success=Compte+refusé+et+supprimé')

@app.route('/admin/supprimer/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_supprimer_utilisateur(id):
    if id == session.get('user_id'):
        return redirect(url_for('admin') + '?tab=users&error=Impossible+de+supprimer+votre+propre+compte')
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=users&success=Utilisateur+supprimé')

@app.route('/admin/changer-role/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_changer_role(id):
    if id == session.get('user_id'):
        return redirect(url_for('admin') + '?tab=users&error=Impossible+de+modifier+votre+propre+rôle')
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT role FROM users WHERE id=?", (id,))
    u = c.fetchone()
    if not u:
        conn.close()
        return redirect(url_for('admin') + '?tab=users&error=Utilisateur+introuvable')
    nouveau_role = 'commercial' if u['role'] == 'admin' else 'admin'
    c.execute("UPDATE users SET role=? WHERE id=?", (nouveau_role, id))
    conn.commit()
    conn.close()
    label = 'promu Administrateur' if nouveau_role == 'admin' else 'rétrogradé Utilisateur'
    return redirect(url_for('admin') + f'?tab=users&success=Utilisateur+{label}')

@app.route('/admin/reset_mdp/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_reset_mdp(id):
    nouveau_mdp = request.form.get('nouveau_mdp', '')
    confirmation_mdp = request.form.get('confirmation_mdp', '')
    if not nouveau_mdp or len(nouveau_mdp) < 6:
        return redirect(url_for('admin') + '?tab=users&error=Mot+de+passe+trop+court')
    if nouveau_mdp != confirmation_mdp:
        return redirect(url_for('admin') + '?tab=users&error=Mots+de+passe+différents')
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT username FROM users WHERE id = ?", (id,))
    u = c.fetchone()
    if not u:
        conn.close()
        return redirect(url_for('admin') + '?tab=users&error=Utilisateur+introuvable')
    c.execute("UPDATE users SET password = ? WHERE id = ?", (generate_password_hash(nouveau_mdp), id))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=users&success=Mot+de+passe+réinitialisé')

@app.route('/admin/reset', methods=['POST'])
@login_required
@admin_required
def admin_reset():
    confirmation = request.form.get('confirmation', '').strip()
    if confirmation != 'RESET':
        return redirect(url_for('admin') + '?tab=danger&error=Confirmation+incorrecte')
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("DELETE FROM interventions")
    c.execute("DELETE FROM historique_statut")
    c.execute("DELETE FROM prospects")
    c.execute("DELETE FROM evenements_perso")
    c.execute("DELETE FROM logs")
    c.execute("DELETE FROM prospect_tags")
    conn.commit()
    conn.close()
    conn2 = sqlite3.connect(DB_NAME)
    c2 = conn2.cursor()
    ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    c2.execute("INSERT INTO logs (username, action, ip) VALUES (?, ?, ?)",
               (session.get('username'), 'REMISE À ZÉRO COMPLÈTE', ip))
    conn2.commit()
    conn2.close()
    return redirect(url_for('admin') + '?tab=danger&success=Remise+à+zéro+effectuée')

@app.route('/admin/backup/download')
@login_required
@admin_required
def admin_backup_download():
    import glob, shutil
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
    # Chercher la sauvegarde la plus récente
    files = sorted(glob.glob(os.path.join(backup_dir, '*.db')), key=os.path.getmtime, reverse=True)
    if files:
        src = files[0]
        fname = os.path.basename(src)
    else:
        # Pas de backup dispo, envoyer la BDD directe
        src = DB_NAME
        fname = f'crm_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db'
    with open(src, 'rb') as f:
        data = f.read()
    return Response(data, mimetype='application/octet-stream',
                    headers={'Content-Disposition': f'attachment; filename={fname}'})

@app.route('/admin/logs')
@login_required
@admin_required
def admin_logs():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM logs ORDER BY horodatage DESC LIMIT 500")
    logs = c.fetchall()
    conn.close()
    return render_template('logs.html', logs=logs)

# ─────────────────────────────────────────────
# FAVICON
# ─────────────────────────────────────────────

@app.route('/admin/upload-favicon', methods=['POST'])
@login_required
@admin_required
def admin_upload_favicon():
    if 'favicon' not in request.files or request.files['favicon'].filename == '':
        return redirect(url_for('admin') + '?tab=perso&error=Aucun+fichier+sélectionné')
    file = request.files['favicon']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in {'png', 'jpg', 'jpeg', 'svg', 'ico', 'gif', 'webp'}:
        return redirect(url_for('admin') + '?tab=perso&error=Format+non+supporté')
    filename = f'favicon_custom.{ext}'
    file.save(os.path.join(app.static_folder, filename))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("INSERT OR REPLACE INTO settings (key, value) VALUES ('favicon_filename', ?)", (filename,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin') + '?tab=perso&success=Favicon+enregistré')

@app.route('/favicon.ico')
def favicon():
    favicon_filename = get_setting('favicon_filename', '')
    if favicon_filename:
        return send_from_directory(app.static_folder, favicon_filename)
    return send_from_directory(app.static_folder, 'favicon.svg')

@app.route('/manifest.json')
def pwa_manifest():
    name = get_setting('company_name', 'OpenSuivi')
    color_primary = get_setting('color_primary', '') or '#0f172a'
    logo_filename = get_setting('logo_filename', '')
    icon_url = f'/static/{logo_filename}' if logo_filename else '/static/favicon.svg'
    data = {
        "name": name,
        "short_name": name,
        "description": "Suivi de recherche d'emploi",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#0f172a",
        "theme_color": color_primary,
        "orientation": "portrait-primary",
        "icons": [
            {"src": icon_url, "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": icon_url, "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}
        ]
    }
    return Response(json.dumps(data), mimetype='application/manifest+json')

@app.route('/sw.js')
def service_worker():
    response = send_from_directory(app.static_folder, 'sw.js', mimetype='application/javascript')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Cache-Control'] = 'no-cache'
    return response

@app.route('/offline')
def offline_page():
    return render_template('offline.html')

# ─────────────────────────────────────────────
# SCHEDULER (backup auto)
# ─────────────────────────────────────────────

def _auto_backup():
    try:
        backup_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backup.py')
        subprocess.run(['python3', backup_script], timeout=60, check=True)
        print(f"[Scheduler] Backup automatique effectué à {datetime.now()}")
    except Exception as e:
        print(f"[Scheduler] Erreur backup auto: {e}")

def _start_scheduler():
    scheduler = BackgroundScheduler()
    scheduler.add_job(_auto_backup, 'cron', hour=2, minute=0, id='backup_auto')
    scheduler.start()
    return scheduler

# ─────────────────────────────────────────────
# LANCEMENT
# ─────────────────────────────────────────────

init_db()
_start_scheduler()

if __name__ == '__main__':
    debug_mode = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port, debug=debug_mode)
